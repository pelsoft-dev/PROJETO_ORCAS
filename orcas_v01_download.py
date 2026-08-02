import io
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from orcas_v01_security import supabase

def render_download(usuario_id=None, projeto_id=None):
    st.subheader("📥 Exportar Lançamentos")
    st.write("Baixe os lançamentos do seu projeto ativo em formato Excel (.xlsx).")

    # --- RESOLUÇÃO FLEXÍVEL DAS CHAVES DE SESSÃO ---
    usr_id = usuario_id or st.session_state.get("user_id") or st.session_state.get("usuario_id") or st.session_state.get("usuario")
    proj_id = projeto_id or st.session_state.get("projeto_ativo") or st.session_state.get("projeto") or st.session_state.get("plano_ativo")

    if not usr_id or not proj_id:
        st.error("⚠️ Usuário ou Projeto Ativo não identificados na sessão. Efetue o login novamente.")
        return

    st.info(f"📌 **Filtro Aplicado:** Usuário `{usr_id}` | Projeto `{proj_id}`")

    if st.button("Gerar Planilha para Download", type="primary", use_container_width=True):
        with st.spinner("Buscando lançamentos no Supabase..."):
            try:
                # Consulta filtrada estritamente pelo usuario_id e projeto_id
                response = (
                    supabase.table("lancamentos")
                    .select("*")
                    .eq("usuario_id", usr_id)
                    .eq("projeto_id", proj_id)
                    .execute()
                )
                data = response.data

                if not data:
                    st.warning("Nenhum lançamento encontrado para este usuário e projeto.")
                    return

                df = pd.DataFrame(data)

                # --- REMOÇÃO DE COLUNAS INTERNAS (SINCRONIZAÇÃO COM UPLOAD) ---
                colunas_para_remover = ["id", "usuario_id", "projeto_id"]
                df = df.drop(columns=[col for col in colunas_para_remover if col in df.columns])

                # Ordenação exata e limpa das colunas de dados
                col_order = [
                    "descricao", "valor", "tipo", "data_vencimento",
                    "realizado", "valor_realizado", "categoria", "recorrente",
                    "valor_plan", "valor_real", "status", "data", "permite_parcial",
                    "usar_media", "complemento_tipo", "complemento_texto", "correcao_freq",
                    "correcao_valor", "id_pai", "parcial_real", "parcial_data", "regra_parcial"
                ]
                
                # Seleciona apenas as colunas existentes
                existing_cols = [c for c in col_order if c in df.columns]
                other_cols = [c for c in df.columns if c not in existing_cols]
                df = df[existing_cols + other_cols]

                # Garantir a conversão correta dos tipos de data no Pandas antes do Excel
                colunas_data = ["data_vencimento", "data", "parcial_data"]
                for col in colunas_data:
                    if col in df.columns:
                        df[col] = pd.to_datetime(df[col], errors="coerce")

                # Exporta para memória RAM (BytesIO) com Formatação via OpenPyXL
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Lancamentos")
                    
                    workbook = writer.book
                    worksheet = writer.sheets["Lancamentos"]

                    # Definições das Regras de Tipo de Coluna
                    colunas_numericas = ["valor", "valor_plan", "valor_real", "valor_realizado", "parcial_real", "correcao_valor"]
                    colunas_booleanas = ["realizado", "recorrente", "permite_parcial", "usar_media"]

                    # Estilos
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                    
                    align_left = Alignment(horizontal="left", vertical="center")
                    align_right = Alignment(horizontal="right", vertical="center")
                    align_center = Alignment(horizontal="center", vertical="center")

                    # 1. Estilização do Cabeçalho (Linha 1)
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = align_center

                    # Mapeamento dos nomes de colunas
                    col_names = [worksheet.cell(row=1, column=col_idx).value for col_idx in range(1, worksheet.max_column + 1)]

                    # 2. Formatação por Célula / Coluna
                    for row_idx in range(2, worksheet.max_row + 1):
                        for col_idx, col_name in enumerate(col_names, start=1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)

                            # --- CAMPOS DE DATA (dd/mm/aaaa - Alinhado ao Centro) ---
                            if col_name in colunas_data:
                                cell.number_format = "DD/MM/YYYY"
                                cell.alignment = align_center

                            # --- CAMPOS NUMÉRICOS (#.##0,00 - Alinhado à Direita) ---
                            elif col_name in colunas_numericas:
                                cell.number_format = "#,##0.00"
                                cell.alignment = align_right

                            # --- CAMPOS BOOLEANOS (Alinhado ao Centro) ---
                            elif col_name in colunas_booleanas:
                                cell.alignment = align_center

                            # --- CAMPOS ALFANUMÉRICOS / TEXTO (Alinhado à Esquerda) ---
                            else:
                                cell.number_format = "@"
                                cell.alignment = align_left

                    # 3. Ajuste Automático da Largura das Colunas
                    for col in worksheet.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            # Formatação de string limpa para medição de tamanho
                            if cell.value is not None:
                                val_str = cell.value.strftime("%d/%m/%Y") if hasattr(cell.value, "strftime") else str(cell.value)
                            else:
                                val_str = ""
                            if len(val_str) > max_len:
                                max_len = len(val_str)
                        
                        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

                buffer.seek(0)

                st.success(f"✅ {len(df)} lançamentos processados com sucesso!")
                
                # Nome dinâmico para o arquivo exportado
                nome_arquivo = f"orcas_{proj_id}_lancamentos.xlsx".lower().replace(" ", "_")

                # Botão nativo para o usuário baixar no navegador
                st.download_button(
                    label="💾 Clique aqui para Baixar o Arquivo Excel",
                    data=buffer,
                    file_name=nome_arquivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"Erro ao exportar lançamentos: {e}")